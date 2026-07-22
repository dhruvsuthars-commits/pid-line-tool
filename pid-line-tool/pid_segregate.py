"""
P&ID Line List Segregator
Parses the LINE column and writes data into the Linelist_reference template (Final sheet).
"""

import re
import os
import json
import pandas as pd
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

# ── Column mapping: parsed field → template column letter in "Final" sheet ──
TEMPLATE_HEADER_ROW = 6
DATA_START_ROW      = 7

LINE_SOURCE_COL = "LINE"

PARSED_COL_MAP = {
    "Fluid Code":      "C",
    "Sequence No":     "D",
    "Line Size (mm)":  "E",
    "Pipe Class":      "F",
    "Insulation":      "G",
}

COL_MAP = {}

ROW_FILL_EVEN = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
ROW_FILL_ODD  = PatternFill("solid", start_color="F2F9F2", end_color="F2F9F2")

THIN        = Side(style="thin", color="000000")
DATA_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ─────────────────────────────────────────────
def parse_line(line_tag: str) -> dict:
    """
    Parse a LINE tag: <FluidCode>-<SequenceNo>-<LineSize>-<PipeClass>[-<Insulation>]
    e.g.  HPS-120816-50-A5-H   or   CWS-120515-50-A1
    """
    result = {
        "Fluid Code":     "",
        "Sequence No":    "",
        "Line Size (mm)": "",
        "Pipe Class":     "",
        "Insulation":     "",
    }
    if not line_tag or str(line_tag).strip() in ("", "nan", "NaN"):
        return result

    parts = str(line_tag).strip().split("-")
    if len(parts) > 0: result["Fluid Code"]     = parts[0]
    if len(parts) > 1: result["Sequence No"]    = parts[1]
    if len(parts) > 2: result["Line Size (mm)"] = parts[2]
    if len(parts) > 3: result["Pipe Class"]     = parts[3]
    if len(parts) > 4: result["Insulation"]     = parts[4]
    return result


# ─────────────────────────────────────────────
def _write_cell(ws, row: int, col_letter: str, value, fill):
    """Write a styled data cell. Writes even empty string — caller decides what to pass."""
    cell = ws.cell(row=row, column=column_index_from_string(col_letter))
    # Normalise pandas/numpy nan to empty string
    str_val = str(value) if value is not None else ""
    cell.value      = "" if str_val in ("nan", "NaN") else (value if value is not None else "")
    cell.fill       = fill
    cell.border     = DATA_BORDER
    cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell.font       = Font(name="Arial", size=9)


def _clear_data_rows(ws):
    """Clear data rows below the header, preserving formatting."""
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row):
        for cell in row:
            cell.value = None


# ─────────────────────────────────────────────
def _apply_mn_rules(parsed_df: pd.DataFrame, configs: list) -> pd.DataFrame:
    """
    Given a list of config dicts:
      [{"fluid_code": "HPS", "pipe_class": "A5",
        "nor_op_pressure": "10", "nor_op_temp": "150"}, ...]

    Returns a DataFrame with columns ["M", "N"] — one row per data row.
    Empty string = no value to write for that row.
    An empty fluid_code or pipe_class in a config means wildcard (match all).
    First matching rule wins.
    """
    n = len(parsed_df)
    m_vals = [""] * n
    n_vals = [""] * n

    if not configs:
        return pd.DataFrame({"M": m_vals, "N": n_vals})

    # Normalise configs once
    rules = []
    for cfg in configs:
        rules.append({
            "fluid_code":      (cfg.get("fluid_code")      or "").strip(),
            "pipe_class":      (cfg.get("pipe_class")       or "").strip(),
            "nor_op_pressure": (cfg.get("nor_op_pressure")  or "").strip(),
            "nor_op_temp":     (cfg.get("nor_op_temp")      or "").strip(),
        })

    fluid_col = parsed_df["Fluid Code"].astype(str).str.strip() if "Fluid Code" in parsed_df.columns else pd.Series([""] * n)
    class_col = parsed_df["Pipe Class"].astype(str).str.strip()  if "Pipe Class"  in parsed_df.columns else pd.Series([""] * n)

    for i in range(n):
        row_fluid = fluid_col.iloc[i]
        row_class = class_col.iloc[i]
        for rule in rules:
            # Empty filter = wildcard; otherwise must match exactly
            fluid_ok = (rule["fluid_code"] == "") or (rule["fluid_code"] == row_fluid)
            class_ok = (rule["pipe_class"] == "")  or (rule["pipe_class"] == row_class)
            if fluid_ok and class_ok:
                m_vals[i] = rule["nor_op_pressure"]
                n_vals[i] = rule["nor_op_temp"]
                break   # first-match wins

    return pd.DataFrame({"M": m_vals, "N": n_vals})


# ─────────────────────────────────────────────
def _write_rows(ws, df: pd.DataFrame, parsed_df: pd.DataFrame, mn_df: pd.DataFrame | None = None):
    """Core writer: fills all data rows into ws starting at DATA_START_ROW."""
    for i in range(len(df)):
        excel_row = DATA_START_ROW + i
        fill      = ROW_FILL_ODD if i % 2 == 0 else ROW_FILL_EVEN

        # Col A — original LINE tag
        line_val = df[LINE_SOURCE_COL].iloc[i] if LINE_SOURCE_COL in df.columns else ""
        _write_cell(ws, excel_row, "A", line_val, fill)

        # Cols C-G — parsed LINE fields
        for field, col_letter in PARSED_COL_MAP.items():
            value = parsed_df[field].iloc[i] if field in parsed_df.columns else ""
            _write_cell(ws, excel_row, col_letter, value, fill)

        # Direct-mapped columns from AutoCAD source
        for src_col, tpl_col in COL_MAP.items():
            match = next((c for c in df.columns if c.strip().upper() == src_col.strip().upper()), None)
            value = df[match].iloc[i] if match else ""
            _write_cell(ws, excel_row, tpl_col, value, fill)

        # Col Q — source file reference (first column of AutoCAD export)
        col_a_val = df.iloc[i, 0] if len(df.columns) >= 1 else ""
        _write_cell(ws, excel_row, "Q", col_a_val, fill)

        # Cols M & N — Nor.Op.Pressure and Nor.Op.Temp
        if mn_df is not None:
            m_val = mn_df["M"].iloc[i]
            n_val = mn_df["N"].iloc[i]
            if m_val != "":
                _write_cell(ws, excel_row, "M", m_val, fill)
            if n_val != "":
                _write_cell(ws, excel_row, "N", n_val, fill)


# ─────────────────────────────────────────────
def write_to_template(template_path: str, df: pd.DataFrame, parsed_df: pd.DataFrame,
                      output_path: str, configs: list = None):
    """Load template, clear data rows, write all data (with optional M&N)."""
    wb = load_workbook(template_path)
    ws = wb["Final"]
    _clear_data_rows(ws)

    mn_df = _apply_mn_rules(parsed_df, configs) if configs else None
    _write_rows(ws, df, parsed_df, mn_df)

    wb.save(output_path)
    print(f"Saved → {output_path}")


# Dynamic column alias mapping for standard template columns
STANDARD_FIELD_ALIASES = {
    "Fluid Code":      ["Fluid Code", "Fluid", "Service", "FluidCode"],
    "Sequence No":     ["Sequence No", "Sequence Number", "Seq No", "SeqNo", "Sequence"],
    "Line Size (mm)":  ["Line Size (mm)", "Line Size", "Size", "LineSize"],
    "Pipe Class":      ["Pipe Class", "Class", "Spec", "PipeClass"],
    "Insulation":      ["Insulation", "Insulation Code", "InsulationCode"],
}

def get_parsed_df_from_input(df: pd.DataFrame) -> pd.DataFrame:
    """
    Extract or generate parsed_df from input DataFrame.
    If input already has parsed columns (from Philosophy AI), use them directly.
    Otherwise fallback to parse_line on LINE column.
    """
    # Check if input already has parsed columns beyond LINE / Source PDF Name
    existing_parsed = {}
    for std_key, aliases in STANDARD_FIELD_ALIASES.items():
        match_col = next((c for c in df.columns if c in aliases or c.strip().lower() in [a.lower() for a in aliases]), None)
        if match_col:
            existing_parsed[std_key] = df[match_col]

    if existing_parsed:
        # Fill missing standard keys with empty strings
        for std_key in PARSED_COL_MAP.keys():
            if std_key not in existing_parsed:
                existing_parsed[std_key] = [""] * len(df)
        parsed_df = pd.DataFrame(existing_parsed)
        # Retain any extra custom dynamic fields (e.g. Area, Unit)
        for col in df.columns:
            if col not in [LINE_SOURCE_COL, "Source PDF Name"] and col not in parsed_df.columns:
                parsed_df[col] = df[col]
        return parsed_df
    else:
        return pd.DataFrame(df[LINE_SOURCE_COL].apply(parse_line).tolist())


# ─────────────────────────────────────────────
def process_file(input_path: str, template_path: str, output_path: str) -> pd.DataFrame:
    """Read input Excel, parse LINE column, export to template. Returns combined df for preview."""
    df = pd.read_excel(input_path)

    line_col = next((c for c in df.columns if c.strip().upper() == LINE_SOURCE_COL.upper()), None)
    if not line_col:
        raise ValueError(f"No '{LINE_SOURCE_COL}' column found. Columns: {list(df.columns)}")
    if line_col != LINE_SOURCE_COL:
        df = df.rename(columns={line_col: LINE_SOURCE_COL})

    parsed_df = get_parsed_df_from_input(df)

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    write_to_template(template_path, df, parsed_df, output_path, configs=None)

    return pd.concat([df.reset_index(drop=True), parsed_df], axis=1)


# ─────────────────────────────────────────────
def export_with_mn_configs(input_path: str, template_path: str,
                           output_path: str, configs: list) -> str:
    """
    Re-export applying M & N values.

    configs: list of dicts —
      [{"fluid_code": "HPS", "pipe_class": "A5",
        "nor_op_pressure": "10", "nor_op_temp": "150"}, ...]

    Empty fluid_code / pipe_class = wildcard (applies to all rows).
    Returns output_path.
    """
    df = pd.read_excel(input_path)

    line_col = next((c for c in df.columns if c.strip().upper() == LINE_SOURCE_COL.upper()), None)
    if not line_col:
        raise ValueError(f"No '{LINE_SOURCE_COL}' column found. Columns: {list(df.columns)}")
    if line_col != LINE_SOURCE_COL:
        df = df.rename(columns={line_col: LINE_SOURCE_COL})

    parsed_df = get_parsed_df_from_input(df)

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    write_to_template(template_path, df, parsed_df, output_path, configs=configs)

    print(f"Saved → {output_path} (with M&N applied, {len(configs)} rule(s))")
    return output_path


# ─────────────────────────────────────────────
def merge_multiple_files(input_paths: list, template_path: str,
                         output_path: str, configs: list = None) -> tuple:
    """
    Merge multiple Excel input files into a single output workbook.
    
    input_paths: list of file paths to process
    template_path: path to Linelist_reference.xlsx template
    output_path: where to save the merged result
    configs: optional M&N rule configs
    
    Returns (output_path, total_rows, merged_df)
    """
    all_dfs = []
    all_parsed = []
    
    # Read and process each input file
    for input_path in input_paths:
        if not os.path.isfile(input_path):
            print(f"Warning: skipping non-existent file {input_path}")
            continue
            
        df = pd.read_excel(input_path)
        
        # Find LINE column (case-insensitive)
        line_col = next((c for c in df.columns if c.strip().upper() == LINE_SOURCE_COL.upper()), None)
        if not line_col:
            print(f"Warning: no '{LINE_SOURCE_COL}' column in {input_path}, skipping")
            continue
        if line_col != LINE_SOURCE_COL:
            df = df.rename(columns={line_col: LINE_SOURCE_COL})
        
        # Parse LINE tags
        parsed_df = pd.DataFrame(df[LINE_SOURCE_COL].apply(parse_line).tolist())
        
        all_dfs.append(df.reset_index(drop=True))
        all_parsed.append(parsed_df.reset_index(drop=True))
    
    if not all_dfs:
        raise ValueError("No valid input files found with LINE column")
    
    # Merge all dataframes
    merged_df = pd.concat(all_dfs, ignore_index=True)
    merged_parsed = pd.concat(all_parsed, ignore_index=True)
    
    # Write to template
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb = load_workbook(template_path)
    ws = wb["Final"]
    _clear_data_rows(ws)
    
    # Apply M&N rules if provided
    mn_df = _apply_mn_rules(merged_parsed, configs) if configs else None
    _write_rows(ws, merged_df, merged_parsed, mn_df)
    
    wb.save(output_path)
    print(f"Saved merged file → {output_path} ({len(merged_df)} total rows)")
    
    return output_path, len(merged_df), pd.concat([merged_df.reset_index(drop=True), merged_parsed], axis=1)



# ─────────────────────────────────────────────
# Legacy wrapper kept for backwards compatibility
# ─────────────────────────────────────────────
def process_file_with_mn(input_path: str, template_path: str,
                         output_path: str, mn_mapping: dict):
    """Thin wrapper around export_with_mn_configs for backwards compat."""
    configs = []
    for key_str, vals in (mn_mapping or {}).items():
        try:
            key = json.loads(key_str)
        except Exception:
            key = {}
        configs.append({
            "fluid_code":      (key.get("fluid_code")  or "").strip(),
            "pipe_class":      (key.get("pipe_class")   or "").strip(),
            "nor_op_pressure": (vals.get("nor_op_pressure") or "").strip(),
            "nor_op_temp":     (vals.get("nor_op_temp")     or "").strip(),
        })
    export_with_mn_configs(input_path, template_path, output_path, configs)


# ─────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    import glob
    import traceback

    script_dir        = os.path.dirname(os.path.abspath(__file__))
    default_input_dir = os.path.join(script_dir, "input")
    default_output_dir= os.path.join(script_dir, "output")
    template_file     = os.path.join(script_dir, "templates_store", "Linelist_reference.xlsx")

    input_file = None
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
    else:
        candidates = (glob.glob(os.path.join(default_input_dir, "*.xlsx")) +
                      glob.glob(os.path.join(default_input_dir, "*.xls")))
        if candidates:
            input_file = candidates[0]
            print(f"No input argument — using: {input_file}")
        else:
            print("Usage: python pid_segregate.py input/yourfile.xlsx")
            sys.exit(1)

    if os.path.isdir(input_file):
        candidates = (glob.glob(os.path.join(input_file, "*.xlsx")) +
                      glob.glob(os.path.join(input_file, "*.xls")))
        if candidates:
            input_file = candidates[0]
        else:
            print(f"No Excel files found in: {input_file}")
            sys.exit(1)

    if not os.path.isfile(input_file):
        print(f"Input file not found: {input_file}"); sys.exit(1)
    if not os.path.isfile(template_file):
        print(f"Template not found: {template_file}"); sys.exit(1)

    output_file = os.path.join(
        default_output_dir,
        f"{os.path.splitext(os.path.basename(input_file))[0]}_Segregated.xlsx"
    )

    try:
        df_result = process_file(input_file, template_file, output_file)
        print(f"\nParsed {len(df_result)} rows.")
        print(df_result[["LINE","Fluid Code","Sequence No","Line Size (mm)","Pipe Class","Insulation"]].head(10).to_string())
    except Exception:
        traceback.print_exc()
        sys.exit(1)
